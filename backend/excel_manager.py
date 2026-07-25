import os
from typing import Dict, Any, List
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_FILE_PATH = os.path.join(os.path.dirname(__file__), "master_marks.xlsx")
SHEET_NAME = "Results"

# Crimson Red Theme Colors
HEADER_BG_COLOR = "C8102E"  # Muthoot Crimson Red
HEADER_FONT_COLOR = "FFFFFF"
SUBHEADER_BG_COLOR = "E63946"

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=HEADER_FONT_COLOR)
DATA_FONT = Font(name="Calibri", size=11)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")

def ensure_master_workbook(file_path: str = EXCEL_FILE_PATH) -> openpyxl.Workbook:
    """Loads existing master workbook or creates a brand new one with exact Muthoot template format."""
    if os.path.exists(file_path):
        try:
            wb = openpyxl.load_workbook(file_path)
            if SHEET_NAME in wb.sheetnames:
                return wb
        except Exception as e:
            print(f"[ExcelManager] Failed to load existing workbook {file_path}: {e}")

    # Create new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Header Row 1: Grouped questions & main titles
    # Col 1: #, Col 2: Student Name, Col 3: Roll No
    # Col 4-5: Q1, Col 6-7: Q2, ..., Col 22-23: Q10
    # Col 24: Marks Secured, Col 25: Max Marks

    ws.merge_cells("A1:A2")
    ws.cell(row=1, column=1, value="#")

    ws.merge_cells("B1:B2")
    ws.cell(row=1, column=2, value="Student Name")

    ws.merge_cells("C1:C2")
    ws.cell(row=1, column=3, value="Roll No")

    col_idx = 4
    for q_num in range(1, 11):
        start_col = get_column_letter(col_idx)
        end_col = get_column_letter(col_idx + 1)
        ws.merge_cells(f"{start_col}1:{end_col}1")
        ws.cell(row=1, column=col_idx, value=f"Q{q_num}")
        
        ws.cell(row=2, column=col_idx, value="a")
        ws.cell(row=2, column=col_idx + 1, value="b")
        col_idx += 2

    ws.merge_cells(f"{get_column_letter(col_idx)}1:{get_column_letter(col_idx)}2")
    ws.cell(row=1, column=col_idx, value="Marks Secured")
    
    col_idx += 1
    ws.merge_cells(f"{get_column_letter(col_idx)}1:{get_column_letter(col_idx)}2")
    ws.cell(row=1, column=col_idx, value="Max Marks")

    # Apply styling to Header Rows 1 & 2
    header_fill = PatternFill(start_color=HEADER_BG_COLOR, end_color=HEADER_BG_COLOR, fill_type="solid")
    subheader_fill = PatternFill(start_color=SUBHEADER_BG_COLOR, end_color=SUBHEADER_BG_COLOR, fill_type="solid")

    for r in range(1, 3):
        for c in range(1, 26):
            cell = ws.cell(row=r, column=c)
            cell.font = HEADER_FONT
            cell.fill = header_fill if r == 1 else subheader_fill
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

    # Set column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 16
    for c in range(4, 24):
        ws.column_dimensions[get_column_letter(c)].width = 7
    ws.column_dimensions['X'].width = 16
    ws.column_dimensions['Y'].width = 14

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20

    wb.save(file_path)
    return wb

def append_booklet_record(data: Dict[str, Any], file_path: str = EXCEL_FILE_PATH) -> int:
    """Appends a single verified booklet record into the master workbook."""
    wb = ensure_master_workbook(file_path)
    ws = wb[SHEET_NAME]

    # Determine next auto-incrementing ID
    next_id = 1
    max_row = ws.max_row
    
    # Check existing rows (starting at row 3)
    if max_row >= 3:
        for r in range(3, max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val is not None:
                try:
                    next_id = max(next_id, int(val) + 1)
                except ValueError:
                    pass

    target_row = max_row + 1 if max_row >= 2 else 3

    # Build row cell values
    student_name = data.get("student_name", "") or ""
    roll_no = data.get("roll_no", "") or ""
    marks_secured = data.get("marks_secured", "") or ""
    max_marks = data.get("max_marks", "") or ""

    questions = data.get("questions", {}) or {}

    row_values = [next_id, student_name, roll_no]

    for q_num in range(1, 11):
        q_data = questions.get(str(q_num), {}) or {}
        val_a = q_data.get("a")
        val_b = q_data.get("b")
        row_values.append("" if val_a is None else str(val_a))
        row_values.append("" if val_b is None else str(val_b))

    row_values.append(marks_secured)
    row_values.append(max_marks)

    # Write row to worksheet
    ws.row_dimensions[target_row].height = 22
    for c_idx, val in enumerate(row_values, start=1):
        cell = ws.cell(row=target_row, column=c_idx, value=val)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = LEFT_ALIGN if c_idx == 2 else CENTER_ALIGN

    wb.save(file_path)
    return next_id

def get_all_records(file_path: str = EXCEL_FILE_PATH) -> List[Dict[str, Any]]:
    """Reads all saved booklet records from the master Excel workbook."""
    if not os.path.exists(file_path):
        return []

    wb = openpyxl.load_workbook(file_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        return []

    ws = wb[SHEET_NAME]
    records = []

    if ws.max_row < 3:
        return []

    for r in range(3, ws.max_row + 1):
        row_id = ws.cell(row=r, column=1).value
        if row_id is None or str(row_id).strip() == "":
            continue

        student_name = ws.cell(row=r, column=2).value or ""
        roll_no = ws.cell(row=r, column=3).value or ""

        questions = {}
        col_idx = 4
        for q_num in range(1, 11):
            val_a = ws.cell(row=r, column=col_idx).value
            val_b = ws.cell(row=r, column=col_idx + 1).value
            questions[str(q_num)] = {
                "a": None if val_a is None or str(val_a) == "" else str(val_a),
                "b": None if val_b is None or str(val_b) == "" else str(val_b),
                "c": None
            }
            col_idx += 2

        marks_secured = ws.cell(row=r, column=col_idx).value or ""
        max_marks = ws.cell(row=r, column=col_idx + 1).value or ""

        records.append({
            "id": row_id,
            "student_name": str(student_name),
            "roll_no": str(roll_no),
            "questions": questions,
            "marks_secured": str(marks_secured),
            "max_marks": str(max_marks)
        })

    return records

def delete_record_by_id(record_id: int, file_path: str = EXCEL_FILE_PATH) -> bool:
    """Deletes a record row from the master workbook by ID and re-numbers remaining rows."""
    if not os.path.exists(file_path):
        return False

    wb = openpyxl.load_workbook(file_path)
    if SHEET_NAME not in wb.sheetnames:
        return False

    ws = wb[SHEET_NAME]
    row_to_delete = None

    for r in range(3, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val is not None and str(val) == str(record_id):
            row_to_delete = r
            break

    if row_to_delete is not None:
        ws.delete_rows(row_to_delete, 1)
        # Re-number ID column
        current_id = 1
        for r in range(3, ws.max_row + 1):
            if ws.cell(row=r, column=1).value is not None:
                ws.cell(row=r, column=1, value=current_id)
                current_id += 1
        wb.save(file_path)
        return True

    return False

def clear_all_records(file_path: str = EXCEL_FILE_PATH) -> bool:
    """Clears all data rows (row 3 onwards) from the master workbook, resetting it to an empty template."""
    if not os.path.exists(file_path):
        ensure_master_workbook(file_path)
        return True

    wb = openpyxl.load_workbook(file_path)
    if SHEET_NAME not in wb.sheetnames:
        ensure_master_workbook(file_path)
        return True

    ws = wb[SHEET_NAME]
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)

    wb.save(file_path)
    return True
